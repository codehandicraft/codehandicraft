# import cv2
# import os
# import re

# def pointDrawingProcess(str):
#     print(f"point drawing {str}")
#     return "ok"


# dd = {"点画":"pointDrawing"}
# print(eval(dd["点画"]+"Process")(dd["点画"]))

# file_path='./sis_out.jpg'

# file_list = ['./sis_out.jpg', './sis_out.jpg', './sis_out.jpg', './sis_out.jpg', './sis_out.jpg', './sis_out.jpg', './sis_out.jpg', './sis_out.jpg', ]


# def str_to_int(input_str):  
#     # 使用正则表达式匹配数字部分  
#     numbers = re.findall(r'\d+', input_str)  
      
#     # 如果找到数字，则返回转换后的整数；否则返回None  
#     if numbers:  
#         return int(numbers[0])  
#     else:  
#         return -1
  
# print(str_to_int('12.3af45.6'))
# print(str_to_int('12'))
# print(str_to_int(''))
# print(str_to_int('ff'))
# print(str_to_int('ff12.3'))
import openpyxl
from openpyxl.styles import PatternFill, Font
import random

def random_color():
    return "{:02X}{:02X}{:02X}".format(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))

# 创建一个新的工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 创建汉字二维数组（100行，100列）
hanzi_array = [['字' for _ in range(100)] for _ in range(100)]

# 设置背景颜色为黑色
bg_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# 为每个单元格设置字体颜色和背景颜色
for row_index, row in enumerate(hanzi_array, start=1):
    for col_index, hanzi in enumerate(row, start=1):
        # 随机生成一个字体颜色
        font_color = Font(color="00" + random_color())

        # 获取单元格
        cell = ws.cell(row=row_index, column=col_index)

        # 设置单元格值
        cell.value = hanzi

        # 设置字体和背景
        cell.font = font_color
        cell.fill = bg_fill

# 保存到一个Excel文件
wb.save("hanzi_excel.xlsx")


