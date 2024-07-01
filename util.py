import re  
import cv2
import os
import numpy as np
import math
import xlwt
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
# height, width = img.shape[:2] 

# 比week_pixel暗的设置为week_pixel
def week_img(path, ori_img, week_pixel):
    dst_img = 255 - ori_img
    _, dst_img = cv2.threshold(dst_img, 255-week_pixel, 255-week_pixel, cv2.THRESH_BINARY)
    dst_img = 255 - dst_img
    print(f"图像淡化成功, {week_pixel=}")
    return imwrite(path, f"_week_{week_pixel}", dst_img)

def get_week_img(ori_img, week_pixel):
    dst_img = 255 - ori_img
    _, dst_img = cv2.threshold(dst_img, 255-week_pixel, 255-week_pixel, cv2.THRESH_BINARY)
    return 255 - dst_img

def under_pixel_to_dst(img, src_pixel, dst_pixel):
    # 灰度图
    if img.ndim != 2:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    for mm in range(img.shape[0]):
        for nn in range(img.shape[1]):
            if img[mm, nn] < src_pixel:
                img[mm, nn] = dst_pixel
    return img

def upwards_pixel_to_dst(img, src_pixel, dst_pixel):
    for mm in range(img.shape[0]):
        for nn in range(img.shape[1]):
            if img[mm, nn] > src_pixel:
                img[mm, nn] = dst_pixel
    return img

def get_square_img(img):
    img = cv2.resize(img, (img.shape[0], img.shape[0]))
    return img

# 获得指定shape的空白图片
# shape: (height, width)
def get_empty_img(shape, one_channel = True):
    empty_img = np.zeros(shape, np.uint8)
    empty_img.fill(255)
    if not one_channel:
        empty_img = cv2.cvtColor(empty_img, cv2.COLOR_GRAY2BGR)  
    return empty_img

def msgOk(msg):
    return [True, msg]

def msgErr(msg, tips=''):
    return [False, msg, tips]

def gen_new_path(path, suffix, file_type=""):
    path_pair = os.path.splitext(path)
    if file_type == "":
        file_type = path_pair[1]
    return path_pair[0] + suffix + file_type

def imwrite(path, suffix, img):
    path_pair = os.path.splitext(path)
    cv2.imwrite(path_pair[0] + suffix + ".jpg", img)
    return path_pair[0] + suffix + ".jpg"
    # cv2.imwrite(path_pair[0] + suffix + path_pair[1], img)
    # return path_pair[0] + suffix + path_pair[1]

def pil_save(path, suffix, img):
    path_pair = os.path.splitext(path)
    img.save(path_pair[0] + suffix + path_pair[1])
    return path_pair[0] + suffix + path_pair[1]

# 裁边，裁剪掉>=src_pixel的部分
def crop_empty(img, src_pixel=255):
    # 灰度图
    gray_img = img
    if img.ndim != 2:
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    for i in range(4):
        for mm in range(gray_img.shape[0]):
            for nn in range(gray_img.shape[1]):
                if gray_img[mm, nn] < src_pixel:
                    if mm > 0:
                        img = img[mm:img.shape[0], :]
                        gray_img = gray_img[mm:gray_img.shape[0], :]
                    break
            else:
                continue
            break
        # 顺时针90度
        img = cv2.flip(cv2.transpose(img), 1)
        gray_img = cv2.flip(cv2.transpose(gray_img), 1)

    return img

def crop_top_empty(img, src_pixel=255):
    # 灰度图
    gray_img = img
    if img.ndim != 2:
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    for i in range(4):
        for mm in range(gray_img.shape[0]):
            if i == 2:
                break
            for nn in range(gray_img.shape[1]):
                if gray_img[mm, nn] < src_pixel:
                    if mm > 0:
                        img = img[mm:img.shape[0], :]
                        gray_img = gray_img[mm:gray_img.shape[0], :]
                    break
            else:
                continue
            break
        # 顺时针90度
        img = cv2.flip(cv2.transpose(img), 1)
        gray_img = cv2.flip(cv2.transpose(gray_img), 1)
        # break

    return img


def crop_white_border(img):  
    gray = img 
    # 转换为灰度图像 
    if img.ndim != 2:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  
      
    # 二值化处理  
    ret, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)  
      
    # 寻找轮廓  
    contours, hierarchy = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  
      
    # 找到最大的轮廓  
    max_contour = max(contours, key=cv2.contourArea)  
      
    # 计算最大轮廓的边界框  
    rect = cv2.boundingRect(max_contour)  
      
    # 绘制边界框  
    # cv2.rectangle(img, (rect[0], rect[1]), (rect[0] + rect[2], rect[1] + rect[3]), (0, 255, 0), 2)  
      
    # 裁剪图像  
    crop_img = img[rect[1]:rect[1]+rect[3], rect[0]:rect[0]+rect[2]]  
      
    return crop_img

# 裁剪图片，达到指定长宽比
def unify_size_by_crop_img(imgs):
    max_h = 0
    max_w = 0
    min_h, min_w = imgs[0].shape[:2]
    for img in imgs:
        h, w= img.shape[:2]
        if h > max_h:
            max_h = h
        if w > max_w:
            max_w = w
        
        if h < min_h:
            min_h = h
        if w < min_w:
            min_w = w
    # print(f"unify_size_by_crop_img max_h={max_h}, max_w={max_w}, {min_h=}")

    for i in range(len(imgs)):
        # 统一高度
        imgs[i] = resize_img(imgs[i], max_h)
        # print(f"unify_size_by_crop_img: unify h: {imgs[i].shape=}")
    
    min_h, min_w = imgs[0].shape[:2]
    for img in imgs:
        h, w= img.shape[:2]
        if h > max_h:
            max_h = h
        if w > max_w:
            max_w = w
        
        if h < min_h:
            min_h = h
        if w < min_w:
            min_w = w
    
    for i in range(len(imgs)):
        img1 = imgs[i]
        img1_new_h = img1.shape[0]
        img1_new_w = min_w
        # 统一宽度
        left = math.ceil((img1.shape[1]-img1_new_w) / 2)
        right = math.floor((img1.shape[1]-img1_new_w) / 2)
        imgs[i] = img1[:, left : img1.shape[1]-right]
        print(f"unify_size_by_crop_img: unify w: {imgs[i].shape=}")


# 统一图片的高度 = 最大高度
def unify_h(imgs):
    max_h = 0
    max_w = 0
    min_h, min_w = imgs[0].shape[:2]
    for img in imgs:
        h, w= img.shape[:2]
        if h > max_h:
            max_h = h
        if w > max_w:
            max_w = w
        
        if h < min_h:
            min_h = h
        if w < min_w:
            min_w = w
    print(f"unify_h: max_h={max_h}, max_w={max_w}, {min_h=}")

    for i in range(len(imgs)):
        imgs[i] = resize_img(imgs[i], max_h)

# 等比例缩放图片，使高度=dst_h
def resize_img(img, dst_h):
    h, w = img.shape[:2]
    return cv2.resize(img, (w * dst_h // h, dst_h))

def resize_img_w(img, dst_w):
    h, w = img.shape[:2]
    return cv2.resize(img, (dst_w, h * dst_w // w))

# 通过填充空白，统一尺寸
def unify_size(imgs):
    unify_h(imgs)
    print("unify size start")
    max_h = 0
    max_w = 0
    for img in imgs:
        h, w= img.shape[:2]
        if h > max_h:
            max_h = h
        if w > max_w:
            max_w = w
    print(f"max_h={max_h}, max_w={max_w}")
    
    for i in range(len(imgs)):
        h, w = imgs[i].shape[:2]
        top = math.ceil((max_h - h) / 2)
        bottom = math.floor((max_h - h) / 2)
        left = math.ceil((max_w - w) / 2)
        right = math.floor((max_w - w) / 2)
        print(f"ori_shape={imgs[i].shape}, h={h}, w={w}, top={top}, bottom={bottom}")

        imgs[i] = cv2.copyMakeBorder(imgs[i], top, bottom, left, right, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        # imgs[i] = cv2.copyMakeBorder(imgs[i], top, bottom, 0, 0, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        print(imgs[i].shape)

    return imgs

# 通过填充上下空白，统一高度
def unify_size_h(imgs):
    print("unify size start")
    max_h = 0
    max_w = 0
    for img in imgs:
        h, w= img.shape[:2]
        if h > max_h:
            max_h = h
        if w > max_w:
            max_w = w
    print(f"max_h={max_h}, max_w={max_w}")
    
    for i in range(len(imgs)):
        h, w = imgs[i].shape[:2]
        top = math.ceil((max_h - h) / 2)
        bottom = math.floor((max_h - h) / 2)
        left = math.ceil((max_w - w) / 2)
        right = math.floor((max_w - w) / 2)
        print(f"ori_shape={imgs[i].shape}, h={h}, w={w}, top={top}, bottom={bottom}")

        # imgs[i] = cv2.copyMakeBorder(imgs[i], top, bottom, left, right, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        imgs[i] = cv2.copyMakeBorder(imgs[i], top, bottom, 0, 0, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        print(imgs[i].shape)

    return imgs

# 填充空白图片，达到指定长宽比
def unify_h_w_ratio_by_fill_blank(img, h, w, color=[255, 255, 255]):
    # 约分
    h,w = get_gcd(h, w)
    print(f"fill_blank: img_shape={img.shape[:2]}, h={h}, w={w}")
    # 最大比例
    ratio = max(img.shape[0] // h, img.shape[1] // w)
    if h * ratio - img.shape[0] < 0 or w * ratio - img.shape[1] < 0:
        ratio += 1
    # 统一长宽比
    new_h = h * ratio
    new_w = w * ratio
    print(f"{ratio=}, {new_h=}, {new_w=}")
    # 填充空白
    top = math.ceil((new_h - img.shape[0]) / 2)
    bottom = math.floor((new_h - img.shape[0]) / 2)
    left = math.ceil((new_w - img.shape[1]) / 2)
    right = math.floor((new_w - img.shape[1]) / 2)
    img = cv2.copyMakeBorder(img, top, bottom, left, right, cv2.BORDER_CONSTANT, value=color)

    return img

# 裁剪图片，达到指定长宽比
def unify_h_w_ratio_by_crop_img(img1, h, w):
    print(f"crop_img: img_shape={img1.shape[:2]}, h={h}, w={w}")
    h,w = get_gcd(h, w)
    print(f"crop_img: img_shape={img1.shape[:2]}, gcd_h={h}, gcd_w={w}")
    img1_new_h = min(img1.shape[0] // h, img1.shape[1] // w) * h
    img1_new_w = min(img1.shape[0] // h, img1.shape[1] // w) * w
    img1 = img1[(img1.shape[0]-img1_new_h)//2: img1.shape[0]-(img1.shape[0]-img1_new_h) //
                2, (img1.shape[1]-img1_new_w)//2: img1.shape[1]-(img1.shape[1]-img1_new_w)//2]
    return img1

def get_gcd(h, w):
    h = int(h * 10)
    w = int(w * 10)
    gcd = math.gcd(h, w)
    h //= gcd
    w //= gcd
    return h, w

def check_directory_exists(directory_path):
    return os.path.exists(directory_path) and os.path.isdir(directory_path)

def create_dir(path):
    try:
        os.makedirs(path)
        print(f"{path}目录创建成功！")
    except FileExistsError:
        print(f"{path}目录已存在！")

def merge_param(input_param, default_param):
    if len(input_param) >= len(default_param):
        return input_param
    for i in range(len(input_param)):
        default_param[i] = input_param[i]
    return default_param

def get_img_list(path_list, min_img_num=2):
    if len(path_list) < min_img_num:
        print(f'input_file_num is too less, will return failed_mail')
        return msgErr('输入图片数量小于2', '请上传2~10张图片；\n图片不要用超大附件发送；\n图片放在附件中')
    img_list = []
    for path in path_list:
        img = cv2.imread(path)
        if img is None:
            print(f"找不到待处理图片, path={path}")
            return msgErr("存在图片无法读取", "可以将图片发到qq重新保存再通过邮件发送过来")
        img_list.append(img)
        print(path, img.shape)
    return img_list

def get_file_size_MB(path):
    return os.path.getsize(path)/2**20
def get_out_file_list(file_path_list):
    if len(file_path_list) == 0:
        return []
    left, right = 0, 0
    sum_size = 0
    out_file_list = []
    MAX_FILE_SIZE_MB = 40
    for file_path in file_path_list:
        file_size = get_file_size_MB(file_path)
        if file_size >= MAX_FILE_SIZE_MB:
            print(f"{file_path=} size={file_size} too big")
            return []
        sum_size += file_size
        if sum_size >= MAX_FILE_SIZE_MB:
            print(f'{left=}, {right=}, {sum_size=:.2f}MB')
            out_file_list.append(file_path_list[left:right])
            left = right
            sum_size = file_size
        right += 1
    print(f'{left=}, {right=}, {sum_size=:.2f}MB')
    out_file_list.append(file_path_list[left:right])
    return out_file_list

def set_xls_color(char_list, img_color, path):

    # 打开现有的Excel文件
    wb = openpyxl.load_workbook(path)
    ws = wb.active  # 假设我们在第一个工作表上工作

   # 遍历单元格，并使用二维数组中的颜色设置字体颜色
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=0):
        for j, cell in enumerate(row, start=0):
            # if cell.value and i < len(color_array) and j < len(color_array[i]):  # 如果单元格有值，并且颜色数组中有对应的颜色
                # 获取对应的RGB颜色值，并转换为十六进制颜色字符串
                color_hex = "{:02X}{:02X}{:02X}".format(img_color[i, j][2], img_color[i, j][1], img_color[i, j][0])
                # 设置单元格的字体颜色
                cell.font = Font(color=color_hex)

    # 保存修改后的工作簿
    wb.save('colored_hanzi_excel.xlsx')


def create_excel_file(hanzi_data, file_index, img_color):
    # 创建一个新的工作簿和sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # 汉字估计的宽度和高度（基于Excel的默认字体大小）
    # 这些值可能需要根据您的实际使用的字体进行调整
    cell_width = 3  # Excel列宽的单位是字符数
    cell_height = 15  # Excel行高的单位是磅
    for col in ws.iter_cols(min_col=1, max_col=len(hanzi_data[0]), max_row=len(hanzi_data)):
        ws.column_dimensions[col[0].column_letter].width = cell_width
    for row in ws.iter_rows(min_row=1, max_row=len(hanzi_data), max_col=len(hanzi_data[0])):
        ws.row_dimensions[row[0].row].height = cell_height

    # 写入数据到sheet
    for i, row_data in enumerate(hanzi_data, start=0):
        for j, hanzi in enumerate(row_data, start=0):
            cell = ws.cell(row=i+1, column=j+1, value=hanzi)
            # 设置字体颜色和背景色
            font_color_rgb = "{:02X}{:02X}{:02X}".format(img_color[i, j][2], img_color[i, j][1], img_color[i, j][0])
            cell.font = Font(color="00" + font_color_rgb)  # 示例中使用红色，你可以根据需要设置不同的RGB值
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存到一个Excel文件
    wb.save(f"hanzi_excel_{file_index}.xlsx")


def save_list_to_xlsxs(char_list, img_color, path):

    # 每20行数据写入一个新的Excel文件
    for file_index, sheet_index in enumerate(range(0, len(char_list), 20)):
        create_excel_file(char_list[sheet_index:sheet_index+20], file_index + 1, img_color[sheet_index:sheet_index+20])


def save_list_to_shard_xlsx(char_list, img_color, path):

    # def random_rgb():
    #     return "{:02X}{:02X}{:02X}".format(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))

    # 创建一个新的工作簿
    wb = Workbook()

    # 汉字二维数组示例
    # char_list = [['字' for _ in range(100)] for _ in range(100)]

    # 汉字估计的宽度和高度（基于Excel的默认字体大小）
    # 这些值可能需要根据您的实际使用的字体进行调整
    cell_width = 3  # Excel列宽的单位是字符数
    cell_height = 15  # Excel行高的单位是磅

    # 创建多个sheets，每个sheet包含20行
    for sheet_index in range(0, len(char_list), 20):
        if sheet_index == 0:
            ws = wb.active
            ws.title = f"Sheet{sheet_index // 20 + 1}"
        else:
            ws = wb.create_sheet(title=f"Sheet{sheet_index // 20 + 1}")

        # 设置列宽度
        for column in ws.iter_cols(1, 100):
            ws.column_dimensions[column[0].column_letter].width = cell_width

        # 写入汉字数组的一部分到当前sheet
        for i in range(sheet_index, min(sheet_index + 20, len(char_list))):
            for j in range(len(char_list[i])):
                # 设置单元格的值、字体颜色和背景色
                cell = ws.cell(row=i - sheet_index + 1, column=j + 1, value=char_list[i][j])
                font_color_rgb = "{:02X}{:02X}{:02X}".format(img_color[i, j][2], img_color[i, j][1], img_color[i, j][0])
                cell.font = Font(color="00" + font_color_rgb)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置行高度
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_col=100, max_row=20):
            for cell in row:
                ws.row_dimensions[cell.row].height = cell_height

    # 保存到一个Excel文件
    wb.save(path)


def save_list_to_xlsx(char_list, img_color, path, sheet_name = 'sheet1'):

    # 创建工作簿
    wb = Workbook()
    # 激活工作表
    ws = wb.active

    # 汉字二维数组示例
    # char_list = [
    #     ['汉', '字', '演', '示'],
    #     ['文', '本', '样', '例'],
    #     ['数', '据', '填', '充']
    # ]

    # 汉字估计的宽度和高度（基于Excel的默认字体大小）
    # 这些值可能需要根据您的实际使用的字体进行调整
    cell_width = 3  # Excel列宽的单位是字符数
    cell_height = 15  # Excel行高的单位是磅

    # 设置行高
    for i in range(1, len(char_list) + 1):
        ws.row_dimensions[i].height = cell_height

    # 设置列宽
    for i in range(1, len(char_list[0]) + 1):
        ws.column_dimensions[get_column_letter(i)].width = cell_width

    # 设置背景颜色为黑色
    bg_color = PatternFill(start_color="00000000", end_color="00000000", fill_type="solid")

    # 写入汉字并设置颜色
    for i, row in enumerate(char_list, start=0):
        for j, hanzi in enumerate(row, start=0):
            # 随机生成一个RGB颜色代码
            font_color_rgb = "{:02X}{:02X}{:02X}".format(img_color[i, j][2], img_color[i, j][1], img_color[i, j][0])
            # 创建一个字体对象，包括字体颜色
            font_color = Font(color="00" + font_color_rgb)
            # 创建单元格
            cell = ws.cell(row=i+1, column=j+1, value=hanzi)
            # 设置单元格字体
            cell.font = font_color
            # 设置单元格背景颜色
            cell.fill = bg_color
            # print(i,j,img_color.shape,img_color[i, j],font_color_rgb)

    # 保存工作簿
    wb.save(path)


def save_list_to_xls(char_list, path, sheet_name = 'sheet1'):
    dicesDrawingXls = xlwt.Workbook(encoding='utf-8')
    sheet=dicesDrawingXls.add_sheet(sheet_name)

    style = xlwt.XFStyle()
    # 设置字体颜色为白色  
    font = xlwt.Font()
    font.colour_index = 9  # 设置字体颜色为白色
    style.font = font
    # 设置填充颜色为黑色  
    pattern = xlwt.Pattern()  
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # 填充类型为纯色填充  
    pattern.pattern_fore_colour = xlwt.Style.colour_map['black']  # 填充颜色为黑色  
    style.pattern = pattern

    for i in range(len(char_list)):
        for j in range(len(char_list[i])):
            if i>255 or j>255:
                break
            # 60表示一个衡量单位，然后再乘以设置的单位数
            sheet.row(i).height_mismatch = True
            sheet.row(i).height = 4 * 60
            # 一个中文等于两个英文等于两个字符，2为字符数，256表示一个衡量单位
            sheet.col(j).width = 2 * 256
            sheet.write(i,j,char_list[i][j],style)
    dicesDrawingXls.save(path)

    print(f"gen xls char result ok, {path=}")
    return path


# 顺时针旋转angle角度，缺失背景白色（255, 255, 255）填充
def rotate_bound_white_bg(image, angle, color=(255, 255, 255)):
    # 判断是否为单通道图像 
    is_single_channel = False 
    if image.ndim == 2:
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)  
        is_single_channel = True
        imwrite("yz2_temp.jpg", '', image)

    # grab the dimensions of the image and then determine the
    # center
    (h, w) = image.shape[:2]
    (cX, cY) = (w // 2, h // 2)
 
    # grab the rotation matrix (applying the negative of the
    # angle to rotate clockwise), then grab the sine and cosine
    # (i.e., the rotation components of the matrix)
    # -angle位置参数为角度参数负值表示顺时针旋转; 1.0位置参数scale是调整尺寸比例（图像缩放参数），建议0.75
    M = cv2.getRotationMatrix2D((cX, cY), -angle, 1.0)
    cos = np.abs(M[0, 0])
    sin = np.abs(M[0, 1])
 
    # compute the new bounding dimensions of the image
    nW = int((h * sin) + (w * cos))
    nH = int((h * cos) + (w * sin))
 
    # adjust the rotation matrix to take into account translation
    M[0, 2] += (nW / 2) - cX
    M[1, 2] += (nH / 2) - cY
 
    # perform the actual rotation and return the image
    # borderValue 缺失背景填充色彩，此处为白色，可自定义
    image = crop_white_border(cv2.warpAffine(image, M, (nW, nH), borderValue=color))
    if is_single_channel:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  

    return image

# 顺时针旋转
def rotate_img(image, angle, fillcolor=255):
    path = 'temp.jpg'
    cv2.imwrite(path, image)
    pil_img = Image.open(path)
    if image.ndim == 2:
        pil_img = pil_img.rotate(-angle, expand=1, fillcolor=fillcolor)
    else:
        pil_img = pil_img.rotate(-angle, expand=1, fillcolor=(255,255,255))
    # pil_img = pil_img.rotate(-90, expand=1, fillcolor=255)
    pil_img.save(path)
    gray_img = cv2.imread(path)
    if image.ndim == 2:
        gray_img = cv2.imread(path, 0)
    # gray_img = crop_white_border(gray_img)
    gray_img = crop_empty(gray_img)
    return gray_img
  
def str_to_int(input_str):  
    # 使用正则表达式匹配数字部分  
    numbers = re.findall(r'\d+', input_str)  
      
    # 如果找到数字，则返回转换后的整数；否则返回None  
    if numbers:  
        return int(numbers[0])  
    else:  
        return -1
  
def cv2AddChineseText(img, text, position=(0, 0), textColor=(255, 255, 255), textSize=30):
    if (isinstance(img, np.ndarray)):  # 判断是否OpenCV图片类型
        img = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
    # 创建一个可以在给定图像上绘图的对象
    draw = ImageDraw.Draw(img)
    # 字体的格式
    # fontStyle = ImageFont.truetype("./charDrawing/simsun.ttc", textSize, encoding="utf-8")
    fontStyle = ImageFont.truetype("/System/Library/Fonts/Hiragino Sans GB.ttc", textSize, encoding="utf-8")
    # fontStyle.set_variation_by_name("bold")
    # 绘制文本
    draw.text(position, text, textColor, font=fontStyle)
    # 转换回OpenCV格式
    return cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)


# 裁剪为指定h:w比例
def crop_img(img1, h, w):
    print(f"crop_img: img_shape={img1.shape[:2]}, h={h}, w={w}")
    h = int(h * 10)
    w = int(w * 10)
    gcd = math.gcd(h, w)
    h //= gcd
    w //= gcd
    img1_new_h = min(img1.shape[0] // h, img1.shape[1] // w) * h
    img1_new_w = min(img1.shape[0] // h, img1.shape[1] // w) * w
    img1 = img1[(img1.shape[0]-img1_new_h)//2: (img1.shape[0]-img1_new_h)//2 + img1_new_h, (img1.shape[1]-img1_new_w)//2: (img1.shape[1]-img1_new_w)//2 + img1_new_w]
    print(f"crop_img: img_shape={img1.shape[:2]}, gcd_h={h}, gcd_w={w}")
    return img1

if __name__ == "__main__":
    path = "./sm.jpg"
    img = cv2.imread(path)
    week_img(path, img, 245)
